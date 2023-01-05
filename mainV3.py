from kivy.app import App
from kivy.uix.button import Button
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.filechooser import FileChooserListView
from kivy.uix.label import Label

import openpyxl


class SortScoresApp(App):
    def build(self):
        # Create a layout to hold the file chooser and the sort button
        layout = BoxLayout(orientation='vertical')

        # Create a file chooser to select the file to sort
        file_chooser = FileChooserListView(path='.')
        layout.add_widget(file_chooser)

        # Create a sort button
        sort_button = Button(text='Sort scores')
        layout.add_widget(sort_button)

        # Bind the sort_scores function to the sort button's on_release event
        sort_button.bind(on_release=lambda instance: self.sort_scores(file_chooser))

        return layout

    def sort_scores(self, file_chooser):
        # Get the selected file from the file chooser
        file_path = file_chooser.selection[0]

        # Open the workbook and sheet
        wb = openpyxl.load_workbook(file_path)
        sheet = wb['Sheet1']

        # Create a list to store the rows
        rows = []

        # Iterate through the rows in the sheet
        for row in sheet.iter_rows():
            # Get the name and score from the row
            name = row[0].value
            score = row[1].value

            # Convert the score to an integer if it is a numeric value
            if isinstance(score, int):
                # The score is already an integer, so we can use it as-is
                pass
            elif isinstance(score, str) and score.isnumeric():
                # The score is a string, so we need to convert it to an integer
                score = int(score)
            else:
                # The score is not an integer or a numeric string, so we set it to 0
                score = 0

            # Add the name and score to the list as a tuple
            rows.append((name, score))

        # Sort the list by the score, from highest to lowest
        sorted_rows = sorted(rows, key=lambda x: x[1], reverse=True)

        # Clear the sheet
        sheet.delete_rows(1, sheet.max_row)

        # Iterate through the sorted rows and add them to the sheet
        for name, score in sorted_rows:
            sheet.append((name, score))

        # Save the workbook
        wb.save(file_path)

        # Show a message to confirm that the scores have been sorted
        self.show_message('Scores sorted successfully!')

    def show_message(self, message):
        # Create a message box
        message_box = BoxLayout(orientation='vertical')

        # Create a label for the message
        label = Label(text=message)
        message_box.add_widget(label)

        # Create a close button
        close_button = Button(text='Close')
        message_box.add_widget(close_button)

        # Bind the close_message function to the close button's on_release event
        close_button.bind(on_release=self.close_message)

        # Add the message box to the app
        self.root.add_widget(message_box)

    def close_message(self, instance):
        # Remove the message box from the app
        self.root.remove_widget(instance.parent)

if __name__ == '__main__':
    SortScoresApp().run()
